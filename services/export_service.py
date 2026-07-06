import os
from datetime import date
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from models import trip as trip_model
from models import payment as payment_model

EXPORT_DIR = os.path.join('static', 'exports')

HEADER_FILL = PatternFill(start_color='D97706', end_color='D97706', fill_type='solid')
HEADER_FONT = Font(bold=True, color='FFFFFF')

STATUS_FILLS = {
    'paid':    (PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid'), Font(color='006100')),
    'partial': (PatternFill(start_color='FFEB9C', end_color='FFEB9C', fill_type='solid'), Font(color='9C6500')),
    'unpaid':  (PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid'), Font(color='9C0006')),
}

INR_FORMAT = '"₹"#,##,##0'

TRIP_HEADERS = ["LR Number", "Date", "Client Name", "From", "To", "Vehicle", "Driver",
                "Weight (MT)", "Packages", "Freight (₹)", "Advance (₹)", "Balance (₹)",
                "Status", "Payment Status"]


def _ensure_export_dir() -> None:
    os.makedirs(EXPORT_DIR, exist_ok=True)


def _style_header_row(ws, row_idx: int = 1) -> None:
    for cell in ws[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center')


def _autofit_columns(ws) -> None:
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[col_cells[0].column_letter].width = min(length + 3, 40)


def _apply_status_fill(cell, payment_status: str) -> None:
    fill, font = STATUS_FILLS.get(payment_status, (None, None))
    if fill:
        cell.fill = fill
        cell.font = font


def _format_date(iso_str: str) -> str:
    if not iso_str:
        return '—'
    return f"{iso_str[8:10]}-{iso_str[5:7]}-{iso_str[:4]}"


def _write_trip_rows(ws, trips: list) -> tuple[float, float, float]:
    total_freight = total_advance = total_balance = 0.0
    for t in trips:
        ws.append([
            t['lr_number'] or '—',
            _format_date(t['created_at']),
            t['client_name'] or '—',
            t['from_location'], t['to_location'],
            t['plate_number'] or '—', t['driver_name'] or '—',
            t['weight_tons'] or 0, t['num_packages'] or 0,
            t['freight_amount'], t['advance_paid'], t['balance_amount'],
            t['status'].replace('_', ' ').title(),
            t['payment_status'].title(),
        ])
        total_freight += t['freight_amount']
        total_advance += t['advance_paid']
        total_balance += t['balance_amount']
        row = ws[ws.max_row]
        for col in (10, 11, 12):
            row[col - 1].number_format = INR_FORMAT
        _apply_status_fill(row[13], t['payment_status'])
    return total_freight, total_advance, total_balance


def export_trips_to_excel(filters: dict | None = None) -> str:
    filters = filters or {}
    trips = trip_model.get_all_trips(
        status_filter=filters.get('status') or None,
        client_filter=int(filters['client_id']) if filters.get('client_id') else None,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Trips"
    ws.append(TRIP_HEADERS)

    total_freight, total_advance, total_balance = _write_trip_rows(ws, trips)

    ws.append(["", "", "", "", "", "", "", "", "TOTAL", total_freight, total_advance, total_balance, "", ""])
    total_row = ws[ws.max_row]
    for cell in total_row:
        cell.font = Font(bold=True)
    for col in (10, 11, 12):
        total_row[col - 1].number_format = INR_FORMAT

    _style_header_row(ws)
    _autofit_columns(ws)

    _ensure_export_dir()
    filename = f"trips_export_{date.today().isoformat()}.xlsx"
    file_path = os.path.join(EXPORT_DIR, filename)
    wb.save(file_path)
    return file_path


def export_dues_to_excel() -> str:
    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Client Dues Summary"
    ws1.append(["Client Name", "Total Trips", "Total Freight (₹)", "Total Received (₹)", "Total Pending (₹)"])

    dues = payment_model.get_client_dues()
    total_trips = total_freight = total_received = total_pending = 0
    for d in dues:
        ws1.append([
            d['client_name'], d['trip_count'],
            d['total_freight'], d['total_received'], d['total_pending'],
        ])
        total_trips += d['trip_count']
        total_freight += d['total_freight']
        total_received += d['total_received']
        total_pending += d['total_pending']
        row = ws1[ws1.max_row]
        for col in (3, 4, 5):
            row[col - 1].number_format = INR_FORMAT

    ws1.append(["TOTAL", total_trips, total_freight, total_received, total_pending])
    total_row = ws1[ws1.max_row]
    for cell in total_row:
        cell.font = Font(bold=True)
    for col in (3, 4, 5):
        total_row[col - 1].number_format = INR_FORMAT

    _style_header_row(ws1)
    _autofit_columns(ws1)

    ws2 = wb.create_sheet("Unpaid Trip Details")
    ws2.append(["LR Number", "Client Name", "From", "To", "Freight (₹)", "Received (₹)", "Pending (₹)", "Payment Status"])

    all_trips = trip_model.get_all_trips()
    unpaid_trips = [t for t in all_trips if t['payment_status'] != 'paid' and t['status'] != 'cancelled']
    for t in unpaid_trips:
        pending = t['freight_amount'] - t['total_received']
        ws2.append([
            t['lr_number'] or '—', t['client_name'] or '—',
            t['from_location'], t['to_location'],
            t['freight_amount'], t['total_received'], pending,
            t['payment_status'].title(),
        ])
        row = ws2[ws2.max_row]
        for col in (5, 6, 7):
            row[col - 1].number_format = INR_FORMAT
        _apply_status_fill(row[7], t['payment_status'])

    _style_header_row(ws2)
    _autofit_columns(ws2)

    _ensure_export_dir()
    filename = f"dues_export_{date.today().isoformat()}.xlsx"
    file_path = os.path.join(EXPORT_DIR, filename)
    wb.save(file_path)
    return file_path


def export_monthly_report(month: int, year: int) -> str:
    trips = trip_model.get_trips_by_month(month, year)
    active_trips = [t for t in trips if t['status'] != 'cancelled']

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Summary"
    total_revenue = sum(t['freight_amount'] for t in active_trips)
    total_pending = sum(t['freight_amount'] - t['total_received'] for t in active_trips)

    ws1.append(["Metric", "Value"])
    ws1.append(["Total Trips", len(active_trips)])
    ws1.append(["Total Revenue (₹)", total_revenue])
    ws1["B3"].number_format = INR_FORMAT
    ws1.append(["Total Pending Dues (₹)", total_pending])
    ws1["B4"].number_format = INR_FORMAT
    ws1.append([])
    ws1.append(["Top Clients", "Freight (₹)"])

    client_totals: dict[str, float] = {}
    for t in active_trips:
        name = t['client_name'] or '—'
        client_totals[name] = client_totals.get(name, 0) + t['freight_amount']
    top_clients = sorted(client_totals.items(), key=lambda x: x[1], reverse=True)[:5]
    for name, amount in top_clients:
        ws1.append([name, amount])
        ws1[f"B{ws1.max_row}"].number_format = INR_FORMAT

    _style_header_row(ws1, 1)
    ws1["A6"].font = Font(bold=True)
    ws1["B6"].font = Font(bold=True)
    _autofit_columns(ws1)

    ws2 = wb.create_sheet("All Trips")
    ws2.append(TRIP_HEADERS)
    total_freight, total_advance, total_balance = _write_trip_rows(ws2, trips)
    ws2.append(["", "", "", "", "", "", "", "", "TOTAL", total_freight, total_advance, total_balance, "", ""])
    total_row = ws2[ws2.max_row]
    for cell in total_row:
        cell.font = Font(bold=True)
    for col in (10, 11, 12):
        total_row[col - 1].number_format = INR_FORMAT
    _style_header_row(ws2)
    _autofit_columns(ws2)

    ws3 = wb.create_sheet("Client-wise Revenue")
    ws3.append(["Client Name", "Trip Count", "Total Freight (₹)"])
    client_counts: dict[str, int] = {}
    for t in active_trips:
        name = t['client_name'] or '—'
        client_counts[name] = client_counts.get(name, 0) + 1
    for name, amount in sorted(client_totals.items(), key=lambda x: x[1], reverse=True):
        ws3.append([name, client_counts.get(name, 0), amount])
        ws3[f"C{ws3.max_row}"].number_format = INR_FORMAT
    _style_header_row(ws3)
    _autofit_columns(ws3)

    _ensure_export_dir()
    filename = f"monthly_report_{year:04d}-{month:02d}.xlsx"
    file_path = os.path.join(EXPORT_DIR, filename)
    wb.save(file_path)
    return file_path
